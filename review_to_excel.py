# review_to_excel.py
import json, argparse, os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import ColorScaleRule

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--review", type=str, default="review_hard.json", help="难样本JSON")
    ap.add_argument("--labelmap", type=str, default="./runs/distilbert_balanced20k/label_mapping.json",
                    help="包含label2id/id2label的映射文件")
    ap.add_argument("--out", type=str, default="review_hard.xlsx")
    args = ap.parse_args()

    # 读取难样本
    with open(args.review, "r", encoding="utf-8") as f:
        data = json.load(f)
    df = pd.DataFrame(data)

    # 尝试读取标签集合
    if os.path.exists(args.labelmap):
        with open(args.labelmap, "r", encoding="utf-8") as f:
            m = json.load(f)
        labels = list(m["label2id"].keys())
    else:
        # 兜底：用文件里出现过的 label / pred_label
        labels = sorted(set(df.get("label", [])) | set(df.get("pred_label", [])))

    # 统一列顺序并新增列
    cols = ["text","label","pred_label","conf","margin","entropy"]
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    df.insert(0, "id", range(1, len(df)+1))
    if "new_label" not in df.columns:
        df["new_label"] = ""

    # 写入Excel
    df = df[["id","text","label","pred_label","conf","margin","entropy","new_label"]]
    df.to_excel(args.out, index=False)

    # 添加下拉&样式
    wb = load_workbook(args.out)
    ws = wb.active
    ws.freeze_panes = "A2"
    # 自动换行 & 列宽
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap
    ws.column_dimensions["A"].width = 6   # id
    ws.column_dimensions["B"].width = 60  # text
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 16  # new_label

    # 下拉（数据验证）在 H 列
    dropdown_vals = labels + ["skip","uncertain"]
    formula = '"' + ",".join(dropdown_vals) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=True)
    dv.error = "请选择列表中的标签"
    dv.prompt = "从下拉中选择修正后的标签"
    ws.add_data_validation(dv)
    dv.add(f"H2:H{ws.max_row}")

    # 条件格式：margin（F列）低值=红色，高值=绿色
    if ws.max_row >= 2:
        rule = ColorScaleRule(start_type='min', start_color='F8696B',
                              mid_type='percentile', mid_value=50, mid_color='FFEB84',
                              end_type='max', end_color='63BE7B')
        ws.conditional_formatting.add(f"F2:F{ws.max_row}", rule)
        # entropy（G列）低值=绿，高值=红（不确定性高）
        rule2 = ColorScaleRule(start_type='min', start_color='63BE7B',
                               mid_type='percentile', mid_value=50, mid_color='FFEB84',
                               end_type='max', end_color='F8696B')
        ws.conditional_formatting.add(f"G2:G{ws.max_row}", rule2)

    wb.save(args.out)
    print(f"✅ 已生成可标注的Excel：{args.out}")
    print(f"📌 标签选项：{dropdown_vals}")

if __name__ == "__main__":
    main()
